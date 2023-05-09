import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Initialize an empty dictionary
data = {}

# Loop through each row and store the data in a dictionary
for row in range(2, worksheet.max_row + 1):
    # Extract the data from the row
    rollno = worksheet.cell(row=row, column=1).value
    name = worksheet.cell(row=row, column=2).value
    marks1 = worksheet.cell(row=row, column=3).value
    marks2 = worksheet.cell(row=row, column=4).value
    marks3 = worksheet.cell(row=row, column=5).value
    
    # Calculate the total marks
    total = marks1 + marks2 + marks3
    
    # Store the data in the dictionary
    data[rollno] = {
        'name': name,
        'marks1': marks1,
        'marks2': marks2,
        'marks3': marks3,
        'total': total
    }

# Display the dictionary data on the monitor
print(data)
